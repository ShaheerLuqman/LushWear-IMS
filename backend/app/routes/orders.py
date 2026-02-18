from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from fastapi.responses import Response
from typing import List, Dict, Any, Optional
from pydantic import BaseModel
from app.models import Order, OrderCreate, OrderUpdate
from app.database import get_supabase
from app.config import settings
from datetime import datetime, timedelta, timezone
import httpx
import re
import csv
import io
import os
from pathlib import Path
from urllib.parse import unquote, urlparse, parse_qs
from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy

router = APIRouter(prefix="/orders", tags=["orders"])

# Pakistan Time (PKT) is UTC+5
PKT_TIMEZONE = timezone(timedelta(hours=5))

def _period_start_end(month: int, year: int):
    """Return (start_iso, end_iso) for period: month's 22 00:00:00 PKT to next month's 22 00:00:00 PKT (exclusive).
    Returns dates in UTC for database comparison.
    PKT (Pakistan Time) is UTC+5, so 00:00 PKT = 19:00 UTC (previous day).
    Example: December period = Dec 22 00:00 PKT (inclusive) to Jan 22 00:00 PKT (exclusive).
    This ensures all of Jan 21 up to 23:59:59.999999 PKT is included."""
    # Create start date: month's 22 at 00:00:00 PKT
    start_pkt = datetime(year, month, 22, 0, 0, 0, 0, tzinfo=PKT_TIMEZONE)
    
    # Calculate next month and year
    next_month = month % 12 + 1
    next_year = year if month != 12 else year + 1
    
    # Create end date: next month's 22 at 00:00:00 PKT (exclusive boundary)
    # This ensures we include everything up to but not including the next period start
    end_pkt = datetime(next_year, next_month, 22, 0, 0, 0, 0, tzinfo=PKT_TIMEZONE)
    
    # Convert to UTC for database comparison
    start_utc = start_pkt.astimezone(timezone.utc)
    end_utc = end_pkt.astimezone(timezone.utc)
    
    # Return ISO format strings with timezone info (Z suffix for UTC)
    return start_utc.isoformat().replace('+00:00', 'Z'), end_utc.isoformat().replace('+00:00', 'Z')


@router.get("/", response_model=List[dict])
async def get_all_orders(
    month: int = Query(None, ge=1, le=12, description="Filter by period month (1-12). Period is 22nd to next 21st."),
    year: int = Query(None, ge=2000, le=2100, description="Filter by period year.")
):
    """Get all orders, optionally filtered by month period (month's 22 to next month's 21)."""
    try:
        supabase = get_supabase()
        if month is not None and year is not None:
            start_iso, end_iso = _period_start_end(month, year)
            # Orders with order_receiving_date in range
            # Start is inclusive (>=), end is exclusive (<) to include all of the last day
            r1 = supabase.table("orders").select("*").gte("order_receiving_date", start_iso).lt("order_receiving_date", end_iso).order("order_number", desc=True).execute()
            # Orders with null order_receiving_date but created_at in range
            r2 = supabase.table("orders").select("*").is_("order_receiving_date", "null").gte("created_at", start_iso).lt("created_at", end_iso).order("order_number", desc=True).execute()
            seen_ids = {o["id"] for o in r1.data}
            merged = list(r1.data)
            for o in r2.data:
                if o["id"] not in seen_ids:
                    merged.append(o)
                    seen_ids.add(o["id"])
            merged.sort(key=lambda x: (x.get("order_number") or 0), reverse=True)
            return merged
        response = supabase.table("orders").select("*").order("order_number", desc=True).execute()
        return response.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/sync-shopify")
async def sync_shopify_orders():
    try:
        store_url = settings.shopify_store_url
        access_token = settings.shopify_access_token
        
        if not store_url or not access_token:
            raise HTTPException(
                status_code=400, 
                detail="Shopify credentials not configured. Please set SHOPIFY_STORE_URL (or SHOPIFY_API_KEY) and SHOPIFY_ADMIN_API_TOKEN environment variables."
            )
            
        store_url = store_url.strip().rstrip('/')
        if store_url.startswith('http://'):
            store_url = store_url[7:]
        elif store_url.startswith('https://'):
            store_url = store_url[8:]
        
        base_url = f"https://{store_url}/admin/api/{settings.SHOPIFY_API_VERSION}/orders.json"
        headers = {
            "X-Shopify-Access-Token": access_token,
            "Content-Type": "application/json"
        }
        
        all_orders = []
        page_info = None
        page_count = 0
        # Only sync orders from the last 30 days
        created_since = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S")

        async with httpx.AsyncClient(timeout=60.0) as client:
            while True:
                if page_info:
                    api_url = f"{base_url}?page_info={page_info}"
                else:
                    api_url = f"{base_url}?status=any&limit=250&created_at_min={created_since}"
                
                response = await client.get(api_url, headers=headers)
                if response.status_code == 404:
                    error_detail = f"Shopify API endpoint not found. Please verify:\n"
                    error_detail += f"1. Store URL is correct: {store_url}\n"
                    error_detail += f"2. API version is valid: {settings.SHOPIFY_API_VERSION}\n"
                    error_detail += f"3. Access token has correct permissions\n"
                    error_detail += f"4. Full URL attempted: {api_url}\n"
                    error_detail += f"Response: {response.text}"
                    raise HTTPException(status_code=404, detail=error_detail)
                response.raise_for_status()
                shopify_data = response.json()
                
                if "orders" not in shopify_data:
                    raise HTTPException(status_code=500, detail="Invalid response from Shopify API")
                
                page_orders = shopify_data["orders"]
                if not page_orders:
                    break
                
                all_orders.extend(page_orders)
                page_count += 1
                
                link_header = response.headers.get("Link", "")
                next_page_info = None
                
                if link_header:
                    next_link_match = re.search(r'<([^>]+)>;\s*rel=["\']next["\']', link_header, re.IGNORECASE)
                    if next_link_match:
                        url = next_link_match.group(1)
                        parsed_url = urlparse(url)
                        if parsed_url.query:
                            query_params = parse_qs(parsed_url.query, keep_blank_values=True)
                            if 'page_info' in query_params:
                                next_page_info = query_params['page_info'][0]
                            else:
                                page_info_match = re.search(r'[?&]page_info=([^&]+)', url)
                                if page_info_match:
                                    next_page_info = unquote(page_info_match.group(1))
                        else:
                            page_info_match = re.search(r'page_info=([^&>]+)', url)
                            if page_info_match:
                                next_page_info = unquote(page_info_match.group(1))
                    
                    if next_page_info:
                        page_info = next_page_info
                        continue
                    else:
                        break
                else:
                    break
                
                if len(page_orders) < 250:
                    break
        
        supabase = get_supabase()
        
        # Fetch all products with their cost prices for cost calculation
        products_response = supabase.table("products").select("name, cost_price").execute()
        products_cost_map = {}
        for p in products_response.data:
            if p.get("name") and p.get("cost_price") is not None:
                # Store by lowercase name for case-insensitive matching
                products_cost_map[p["name"].lower().strip()] = float(p["cost_price"])
        
        existing_orders_map = {}
        existing_orders_all = []
        offset = 0
        limit = 1000
        existing_orders_select = (
            "id, order_number, order_status, delivery_charge, tax_amount, "
            "delivery_status, piece_received, courier, tracking_number, "
            "cost_price, items, total_amount, advance_amount, order_receiving_date"
        )
        while True:
            existing_orders_response = (
                supabase.table("orders")
                .select(existing_orders_select)
                .range(offset, offset + limit - 1)
                .execute()
            )
            if not existing_orders_response.data:
                break
            existing_orders_all.extend(existing_orders_response.data)
            if len(existing_orders_response.data) < limit:
                break
            offset += limit
        
        for o in existing_orders_all:
            order_num = o.get("order_number")
            if order_num is not None:
                try:
                    existing_orders_map[int(order_num)] = o
                except (ValueError, TypeError):
                    continue
        
        def extract_courier(order):
            """Extract courier from the latest non-cancelled fulfillment, or latest fulfillment if all are cancelled."""
            if "fulfillments" not in order or len(order["fulfillments"]) == 0:
                return "Unassigned"
            
            fulfillments = order["fulfillments"]
            
            # Filter out cancelled fulfillments first, but keep them as fallback
            active_fulfillments = [f for f in fulfillments if f.get("status") != "cancelled"]
            fulfillments_to_check = active_fulfillments if active_fulfillments else fulfillments
            
            if not fulfillments_to_check:
                return "Unassigned"
            
            # Find the latest fulfillment by updated_at (or created_at if updated_at is missing)
            latest_fulfillment = None
            latest_timestamp = None
            
            for fulfillment in fulfillments_to_check:
                # Prefer updated_at as it reflects the latest change
                timestamp_str = fulfillment.get("updated_at") or fulfillment.get("created_at")
                if not timestamp_str:
                    continue
                
                timestamp = _parse_iso(timestamp_str)
                if timestamp and (latest_timestamp is None or timestamp > latest_timestamp):
                    latest_timestamp = timestamp
                    latest_fulfillment = fulfillment
            
            # If we couldn't find by timestamp, use the last one in the list
            if not latest_fulfillment:
                latest_fulfillment = fulfillments_to_check[-1]
            
            tracking_company = latest_fulfillment.get("tracking_company")
            if tracking_company:
                tracking_company = str(tracking_company).strip()
                if tracking_company:
                    return tracking_company
            return "Unassigned"
        
        def extract_tracking_number(order):
            """Extract tracking number from the latest non-cancelled fulfillment, or latest fulfillment if all are cancelled."""
            if "fulfillments" not in order or len(order["fulfillments"]) == 0:
                return None
            
            fulfillments = order["fulfillments"]
            
            # Filter out cancelled fulfillments first, but keep them as fallback
            active_fulfillments = [f for f in fulfillments if f.get("status") != "cancelled"]
            fulfillments_to_check = active_fulfillments if active_fulfillments else fulfillments
            
            if not fulfillments_to_check:
                return None
            
            # Find the latest fulfillment by updated_at (or created_at if updated_at is missing)
            latest_fulfillment = None
            latest_timestamp = None
            
            for fulfillment in fulfillments_to_check:
                # Prefer updated_at as it reflects the latest change
                timestamp_str = fulfillment.get("updated_at") or fulfillment.get("created_at")
                if not timestamp_str:
                    continue
                
                timestamp = _parse_iso(timestamp_str)
                if timestamp and (latest_timestamp is None or timestamp > latest_timestamp):
                    latest_timestamp = timestamp
                    latest_fulfillment = fulfillment
            
            # If we couldn't find by timestamp, use the last one in the list
            if not latest_fulfillment:
                latest_fulfillment = fulfillments_to_check[-1]
            
            tracking_number = latest_fulfillment.get("tracking_number")
            if tracking_number:
                tracking_number = str(tracking_number).strip()
                if tracking_number:
                    return tracking_number
            return None
        
        def _parse_iso(s):
            if not s:
                return None
            if isinstance(s, datetime):
                return s
            s = str(s).strip().replace("Z", "+00:00")
            try:
                return datetime.fromisoformat(s)
            except (ValueError, TypeError):
                return None

        def extract_order_status(order):
            cancelled_at_raw = order.get("cancelled_at")
            fulfillment_dt = None
            for f in order.get("fulfillments") or []:
                ct = f.get("created_at")
                if ct:
                    parsed = _parse_iso(ct)
                    if parsed and (fulfillment_dt is None or parsed > fulfillment_dt):
                        fulfillment_dt = parsed
            if cancelled_at_raw and fulfillment_dt is not None:
                cancelled_at = _parse_iso(cancelled_at_raw)
                if cancelled_at and cancelled_at > fulfillment_dt:
                    return "returned"
            if cancelled_at_raw is not None:
                return "cancelled"
            fulfillment_status = order.get("fulfillment_status")
            if fulfillment_status == "fulfilled":
                return "fulfilled"
            return "unfulfilled"
        
        def extract_delivery_status(order):
            fulfillment_status = order.get("fulfillment_status")
            if fulfillment_status == "fulfilled":
                return "delivered"
            elif fulfillment_status == "partial":
                return "partially_delivered"
            elif fulfillment_status is None:
                return "not_delivered"
            else:
                return fulfillment_status or "not_delivered"
        
        def extract_advance_amount(order):
            if "note_attributes" in order:
                for attr in order["note_attributes"]:
                    if attr.get("name") in ["advance", "Advance", "advance_amount"]:
                        try:
                            return float(attr.get("value", 0))
                        except:
                            return None
            return None
        
        def extract_tax_amount(order):
            # Prefer current_* (reflects edits/refunds; avoids discrepancies when order is updated)
            if "current_total_tax_set" in order and order["current_total_tax_set"]:
                shop_money = order["current_total_tax_set"].get("shop_money", {})
                if shop_money:
                    return float(shop_money.get("amount", "0.00"))
            try:
                return float(order.get("current_total_tax") or 0)
            except (TypeError, ValueError):
                pass
            if "total_tax_set" in order and order["total_tax_set"]:
                shop_money = order["total_tax_set"].get("shop_money", {})
                if shop_money:
                    return float(shop_money.get("amount", "0.00"))
            return float(order.get("total_tax", "0.00"))
        
        def extract_cost_price(order):
            if "note_attributes" in order:
                for attr in order["note_attributes"]:
                    if attr.get("name") in ["cost_price", "Cost Price", "cost"]:
                        try:
                            return float(attr.get("value", 0))
                        except:
                            return None
            return None
        
        def calculate_cost_from_items(items, products_cost_map):
            """Calculate total cost price by looking up each item in the products table"""
            if not items:
                return 0.0
            
            total_cost = 0.0
            for item_name in items:
                item_lower = item_name.lower().strip()
                
                # Try exact match first
                if item_lower in products_cost_map:
                    total_cost += products_cost_map[item_lower]
                    continue
                
                # Try matching without variant suffix (e.g., "Product Name - S" -> "Product Name")
                # Items from Shopify are like "Product Name - Variant"
                if " - " in item_name:
                    product_name = item_name.rsplit(" - ", 1)[0].lower().strip()
                    if product_name in products_cost_map:
                        total_cost += products_cost_map[product_name]
                        continue
                
                # Try partial match - find products whose name is contained in item name
                for product_name, cost in products_cost_map.items():
                    if product_name in item_lower or item_lower in product_name:
                        total_cost += cost
                        break
            
            return total_cost
        
        def extract_items(order):
            if "line_items" not in order or not order["line_items"]:
                return []
            item_names = []
            for item in order["line_items"]:
                name = item.get("name", "")
                if name:
                    item_names.append(name)
            return item_names

        def subtotal_line_items_excluding_removed(order):
            """Sum line item totals excluding removed items. Uses current_quantity (Shopify's quantity after removals) when present, else quantity. Removed lines have current_quantity=0."""
            if "line_items" not in order or not order["line_items"]:
                return None
            total = 0.0
            for item in order["line_items"]:
                # current_quantity is the quantity after edits/removals; when a line is removed it is 0
                qty = item.get("current_quantity")
                if qty is None:
                    qty = item.get("quantity") or 0
                try:
                    qty = int(qty)
                except (TypeError, ValueError):
                    qty = 0
                if qty <= 0:
                    continue
                try:
                    price = float(item.get("price") or 0)
                    total += price * qty
                except (TypeError, ValueError):
                    pass
            return total if total > 0 else None
        
        def normalize_value(val):
            if val is None:
                return None
            if isinstance(val, (int, float)):
                return round(float(val), 2)
            return str(val).strip() if val else None
        
        def has_changed(shopify_data, existing_data, skip_assigned_courier_fields=False):
            """
            skip_assigned_courier_fields: when True, do not compare courier, tracking_number,
            total_amount, delivery_charge, tax_amount, cost_price, items (used when courier is assigned).
            """
            fields_to_compare = ["courier", "tracking_number", "order_status", "total_amount", "advance_amount", "delivery_charge", "tax_amount", "cost_price", "items"]
            if skip_assigned_courier_fields:
                fields_to_compare = ["order_status", "piece_received", "advance_amount"]
            for field in fields_to_compare:
                shopify_val = normalize_value(shopify_data.get(field))
                existing_val = normalize_value(existing_data.get(field))
                if field in ["total_amount", "advance_amount", "delivery_charge", "tax_amount", "cost_price"]:
                    shopify_num = float(shopify_val) if shopify_val is not None else 0.0
                    existing_num = float(existing_val) if existing_val is not None else 0.0
                    if abs(shopify_num - existing_num) > 0.01:
                        return True
                elif field == "items":
                    shopify_list = shopify_val if isinstance(shopify_val, list) else []
                    existing_list = existing_val if isinstance(existing_val, list) else []
                    if set(shopify_list) != set(existing_list):
                        return True
                elif field == "courier":
                    shopify_str = (shopify_val or "").strip() or "Unassigned"
                    existing_str = (existing_val or "").strip() or "Unassigned"
                    if shopify_str.lower() != existing_str.lower():
                        return True
                elif field == "tracking_number":
                    shopify_str = (shopify_val or "").strip() if shopify_val else None
                    existing_str = (existing_val or "").strip() if existing_val else None
                    if shopify_str != existing_str:
                        return True
                else:
                    if shopify_val != existing_val:
                        return True
            return False
        
        orders_to_insert = []
        orders_to_update = []
        orders_to_skip = []
        current_time = datetime.utcnow().isoformat()

        for sp_order in all_orders:
            order_number = sp_order.get("order_number")
            if not order_number:
                continue
            
            order_number = int(order_number)
            
            courier = extract_courier(sp_order)
            tracking_number = extract_tracking_number(sp_order)
            order_status = extract_order_status(sp_order)
            
            # Calculate total amount: use only non-removed line items (exclude quantity 0 / removed products)
            total_line_items_price = subtotal_line_items_excluding_removed(sp_order)
            if total_line_items_price is None:
                total_line_items_price = float(sp_order.get("total_line_items_price") or 0)
            shopify_tax = extract_tax_amount(sp_order) or 0.0  # Used only for total_amount calc; we never store tax from Shopify

            # Get shipping price (used only for total_amount calc; we never store delivery_charge from Shopify)
            shipping_price = 0.0
            if "total_shipping_price_set" in sp_order and sp_order["total_shipping_price_set"]:
                shop_money = sp_order["total_shipping_price_set"].get("shop_money", {})
                if shop_money:
                    shipping_price = float(shop_money.get("amount", "0.00"))
            elif "total_shipping_price" in sp_order:
                shipping_price = float(sp_order.get("total_shipping_price") or 0)

            # Total = line items (excluding removed) + tax + shipping
            total_amount = total_line_items_price + shopify_tax + shipping_price
            
            # Sum of all discounts (custom order-level + per-item discounts)
            total_discounts = float(sp_order.get("current_total_discounts") or sp_order.get("total_discounts") or 0)
            
            # Advance: if paid = full order price; if not paid = sum of discounts
            financial_status = (sp_order.get("financial_status") or "").strip().lower()
            if financial_status == "paid":
                advance_amount = total_amount
            else:
                advance_amount = total_discounts
            
            # delivery_charge and tax_amount are never taken from Shopify; set manually or via CSV
            delivery_charge = 0.0
            tax_amount = 0.0
            cost_price = extract_cost_price(sp_order) or 0.0
            
            # Set fixed delivery charge for SCS courier
            if courier.upper() == "SCS":
                delivery_charge = 180.0
            items = extract_items(sp_order)
            
            # If cost_price is 0, calculate it from items using products table
            if cost_price == 0.0 and items:
                cost_price = calculate_cost_from_items(items, products_cost_map)
            
            order_received_date = sp_order.get("created_at")
            if order_received_date:
                try:
                    order_received_date = datetime.fromisoformat(order_received_date.replace('Z', '+00:00')).isoformat()
                except:
                    try:
                        order_received_date = datetime.strptime(order_received_date, "%Y-%m-%dT%H:%M:%S%z").isoformat()
                    except:
                        order_received_date = current_time
            else:
                order_received_date = current_time
            
            order_data = {
                "order_number": order_number,
                "courier": courier,
                "tracking_number": tracking_number,
                "order_status": order_status,
                "piece_received": "Pending",
                "total_amount": total_amount,
                "advance_amount": advance_amount,
                "delivery_charge": delivery_charge,
                "tax_amount": tax_amount,
                "cost_price": cost_price,
                "order_receiving_date": order_received_date,
                "items": items,
                "updated_at": current_time
            }
            
            if order_number in existing_orders_map:
                existing_order = existing_orders_map[order_number]
                existing_status = (existing_order.get("order_status") or "").strip().lower()
                if existing_status in ("delivered", "returned"):
                    existing_adv = float(existing_order.get("advance_amount") or 0)
                    existing_tot = float(existing_order.get("total_amount") or 0)
                    adv_changed = abs(existing_adv - advance_amount) > 0.01
                    tot_changed = abs(existing_tot - total_amount) > 0.01
                    if adv_changed or tot_changed:
                        financial_only = {
                            **existing_order, 
                            "advance_amount": advance_amount, 
                            "total_amount": total_amount,
                            "updated_at": current_time
                        }
                        orders_to_update.append(financial_only)
                    else:
                        orders_to_skip.append(order_number)
                    continue
                shopify_order_status = (order_data.get("order_status") or "").strip().lower()
                if shopify_order_status == "cancelled":
                    order_data["order_status"] = shopify_order_status
                elif shopify_order_status == "fulfilled" and existing_status == "unfulfilled":
                    order_data["order_status"] = shopify_order_status
                else:
                    order_data["order_status"] = existing_order.get("order_status")
                # Advance is always from Shopify: paid = total_amount, not paid = total_discounts
                order_data["advance_amount"] = advance_amount
                # Preserve tax_amount (never from Shopify; set manually or via CSV)
                order_data["tax_amount"] = existing_order.get("tax_amount", 0)
                # Preserve last fetched delivery status (from courier tracking)
                order_data["delivery_status"] = existing_order.get("delivery_status")
                # Preserve piece_received (set by delivery status or manually; default Pending)
                order_data["piece_received"] = existing_order.get("piece_received") or "Pending"

                existing_courier = (existing_order.get("courier") or "").strip()
                existing_tracking = (existing_order.get("tracking_number") or "").strip() if existing_order.get("tracking_number") else None
                courier_is_assigned = bool(existing_courier and existing_courier.lower() != "unassigned")

                # Compare and update courier and tracking_number from Shopify if they differ
                shopify_courier = (courier or "").strip()
                shopify_tracking = (tracking_number or "").strip() if tracking_number else None
                
                # Normalize for comparison (handle "Unassigned" vs empty)
                existing_courier_normalized = existing_courier.lower() if existing_courier else "unassigned"
                shopify_courier_normalized = shopify_courier.lower() if shopify_courier else "unassigned"
                
                # Update courier and tracking_number from Shopify if they differ
                courier_changed = existing_courier_normalized != shopify_courier_normalized
                tracking_changed = existing_tracking != shopify_tracking
                
                if courier_changed or tracking_changed:
                    # Update courier and tracking_number from Shopify
                    order_data["courier"] = courier
                    order_data["tracking_number"] = tracking_number
                else:
                    # Keep existing values if they match
                    order_data["courier"] = existing_order.get("courier")
                    order_data["tracking_number"] = existing_order.get("tracking_number")
                
                # Set delivery_charge: 180 for SCS courier only if not already set to a non-zero value
                # Preserve any manually set delivery_charge (never from Shopify; set manually or via CSV)
                final_courier = (order_data.get("courier") or "").strip().upper()
                existing_delivery_charge = float(existing_order.get("delivery_charge") or 0)
                if final_courier == "SCS" and existing_delivery_charge == 0:
                    # Only set to 180 if courier is SCS and delivery_charge hasn't been set yet
                    order_data["delivery_charge"] = 180.0
                else:
                    # Preserve existing delivery_charge (including any non-zero values)
                    order_data["delivery_charge"] = existing_delivery_charge

                # Preserve existing order_receiving_date - never overwrite from Shopify for existing orders
                order_data["order_receiving_date"] = existing_order.get("order_receiving_date")

                if courier_is_assigned:
                    order_data["total_amount"] = total_amount
                    order_data["advance_amount"] = advance_amount
                    # delivery_charge already set above (180 for SCS, otherwise existing)
                    order_data["tax_amount"] = existing_order.get("tax_amount", 0)
                    order_data["cost_price"] = existing_order.get("cost_price")
                    order_data["items"] = existing_order.get("items")
                    skip_fields = True
                else:
                    skip_fields = False

                # Always update if courier or tracking_number changed, otherwise check other fields
                if courier_changed or tracking_changed or has_changed(order_data, existing_order, skip_assigned_courier_fields=skip_fields):
                    order_data["id"] = existing_order["id"]
                    orders_to_update.append(order_data)
                else:
                    orders_to_skip.append(order_number)
            else:
                # First-time create: sync all fields from Shopify
                order_data["created_at"] = current_time
                orders_to_insert.append(order_data)
        
        created_count = 0
        if orders_to_insert:
            batch_size = 1000
            for i in range(0, len(orders_to_insert), batch_size):
                batch = orders_to_insert[i:i + batch_size]
                supabase.table("orders").upsert(batch, on_conflict="order_number").execute()
                created_count += len(batch)
        
        updated_count = 0
        if orders_to_update:
            batch_size = 1000
            for i in range(0, len(orders_to_update), batch_size):
                batch = orders_to_update[i:i + batch_size]
                supabase.table("orders").upsert(batch, on_conflict="order_number").execute()
                updated_count += len(batch)
        
        synced_count = created_count + updated_count
        skipped_count = len(orders_to_skip)

        return {
            "message": "Orders synced successfully",
            "synced": synced_count,
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "pages_fetched": page_count,
            "total_orders_from_shopify": len(all_orders),
            "orders_per_page": 250 if len(all_orders) > 0 else 0
        }
        
    except HTTPException:
        raise
    except httpx.HTTPStatusError as e:
        error_text = e.response.text
        try:
            error_json = e.response.json()
            error_text = str(error_json)
        except:
            pass
        raise HTTPException(
            status_code=e.response.status_code,
            detail=f"Shopify API error: {error_text}\nURL: {api_url if 'api_url' in locals() else 'N/A'}"
        )
    except httpx.RequestError as e:
        raise HTTPException(status_code=500, detail=f"Failed to connect to Shopify: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error syncing orders: {str(e)}")


def _parse_float(val: Any, default: float = 0.0) -> float:
    """Parse a CSV cell to float; return default if invalid. Strips commas and trailing %."""
    if val is None or (isinstance(val, str) and val.strip() == ""):
        return default
    try:
        s = str(val).strip().replace(",", "").rstrip("%").strip()
        return float(s) if s else default
    except (ValueError, TypeError):
        return default


@router.post("/upload-postex-csv")
async def upload_postex_csv(file: UploadFile = File(...)):
    """
    Upload a PostEx CSV file. Matches rows by ORDER_REF_NUMBER to orders and updates
    delivery_charge (from SHIPPING_CHARGES), tax_amount (GST + WH_INCOME_TAX + WH_SALES_TAX),
    courier (set to PostEx), and tracking_number (from TRACKING_NUMBER; parses 14-digit numbers
    including exponential notation e.g. 2.63E+13).
    """
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Please upload a CSV file.")
    try:
        content = await file.read()
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV has no header row.")
        # Map possible column names to canonical keys
        # Use original fieldnames (not stripped) because DictReader uses original fieldnames as keys
        col_map = {}
        for i, name in enumerate(reader.fieldnames):
            key_upper = name.upper().strip()  # Compare against stripped uppercase
            key_norm = key_upper.replace(" ", "_")  # "WH INCOME TAX (2%)" -> "WH_INCOME_TAX_(2%)"
            # Check each pattern independently (not elif) to handle multiple columns
            # Store the ORIGINAL fieldname (with spaces if any) for use with row.get()
            if "SHIPPING_CHARGES" in key_upper and "shipping_charges" not in col_map:
                col_map["shipping_charges"] = name  # Use original name
            if "GST" in key_upper and "TAX" not in key_upper and "gst" not in col_map:
                col_map["gst"] = name  # Use original name
            # WH_INCOME_TAX (2%) or WH INCOME TAX (2%) or INCOME_TAX
            if ("wh_income_tax" not in col_map and (
                "WH_INCOME_TAX" in key_norm or "INCOME_TAX" in key_norm
                or ("WH" in key_upper and "INCOME" in key_upper and "TAX" in key_upper and "SALES" not in key_upper)
            )):
                col_map["wh_income_tax"] = name
            # WH_SALES_TAX (2%) or WH SALES TAX (2%) or SALES_TAX
            if ("wh_sales_tax" not in col_map and (
                "WH_SALES_TAX" in key_norm or "SALES_TAX" in key_norm
                or ("WH" in key_upper and "SALES" in key_upper and "TAX" in key_upper)
            )):
                col_map["wh_sales_tax"] = name
            if ("ORDER_REF_NUMBER" in key_upper or "ORDER_NUMBER" in key_upper or "ORDER_ID" in key_upper) and "order_ref_number" not in col_map:
                col_map["order_ref_number"] = name  # Use original name
            if ("TRACKING_NUMBER" in key_upper or "TRACKING" in key_upper) and "tracking_number" not in col_map:
                col_map["tracking_number"] = name
            if "NET_AMOUNT" in key_upper and "net_amount" not in col_map:
                col_map["net_amount"] = name
        if "order_ref_number" not in col_map:
            raise HTTPException(status_code=400, detail="CSV must contain an ORDER_REF_NUMBER, ORDER_NUMBER, or ORDER_ID column.")
        if "shipping_charges" not in col_map:
            raise HTTPException(status_code=400, detail="CSV must contain SHIPPING_CHARGES column.")
        tracking_col = col_map.get("tracking_number")
        
        def normalize_order_number(order_ref):
            """Extract order number from formats like #4807 or 4446-R using regex (first run of digits)."""
            if order_ref is None:
                return None
            if isinstance(order_ref, (int, float)):
                return int(order_ref)
            order_str = str(order_ref).strip()
            if not order_str:
                return None
            match = re.search(r"\d+", order_str)
            if not match:
                return None
            try:
                return int(match.group(0))
            except (ValueError, TypeError):
                return None

        def parse_tracking_number_14(val):
            """Parse 14-digit tracking number; CSV may show it as exponential (e.g. 2.63E+13)."""
            if val is None:
                return None
            s = str(val).strip()
            if not s:
                return None
            try:
                # Handle exponential notation (e.g. 2.63E+13 -> 26300000000000)
                if "e" in s.lower():
                    n = int(float(s))
                else:
                    n = int(s)
                # Return as 14-digit string (zero-pad if needed)
                return str(n).zfill(14) if 0 <= n < 10**14 else str(n)
            except (ValueError, TypeError):
                return None

        rows = []
        csv_order_numbers = []
        for row in reader:
            # Get values using the actual fieldnames from col_map
            order_ref_raw = row.get(col_map["order_ref_number"], "")
            order_number = normalize_order_number(order_ref_raw)
            if not order_number:
                continue
            shipping = _parse_float(row.get(col_map["shipping_charges"], ""), 0)
            gst = _parse_float(row.get(col_map.get("gst", ""), ""), 0)
            income_tax = _parse_float(row.get(col_map.get("wh_income_tax", ""), ""), 0)
            sales_tax = _parse_float(row.get(col_map.get("wh_sales_tax", ""), ""), 0)
            shipping_total = shipping + gst
            tax_total = income_tax + sales_tax
            tracking_raw = row.get(tracking_col, "") if tracking_col else ""
            tracking_str = parse_tracking_number_14(tracking_raw)
            net_amount_raw = row.get(col_map.get("net_amount", ""), "") if col_map.get("net_amount") else None
            net_amount_val = _parse_float(net_amount_raw, None) if net_amount_raw is not None and str(net_amount_raw).strip() != "" else None
            rows.append({
                "order_number": order_number,
                "delivery_charge": shipping_total,
                "tax_amount": tax_total,
                "tracking_number": tracking_str,
                "csv_net_amount": net_amount_val,
            })
            csv_order_numbers.append(order_number)
        if not rows:
            return {"updated": 0, "message": "No valid rows with ORDER_REF_NUMBER in CSV."}
        supabase = get_supabase()
        # Fetch all orders (we need to match by order_number)
        all_orders = []
        limit = 1000
        offset = 0
        while True:
            resp = supabase.table("orders").select("id, order_number, total_amount, advance_amount, order_status").range(offset, offset + limit - 1).execute()
            if not resp.data:
                break
            all_orders.extend(resp.data)
            if len(resp.data) < limit:
                break
            offset += limit
        order_number_to_order = {}
        db_order_numbers = []
        for o in all_orders:
            on = o.get("order_number")
            if on is not None:
                try:
                    order_num = int(on)
                    order_number_to_order[order_num] = o
                    db_order_numbers.append(order_num)
                except (ValueError, TypeError):
                    continue
        
        # Find matches and detect receivable vs CSV net amount mismatches
        matched_order_numbers = []
        updated_count = 0
        unmatched_order_numbers = []
        updated_order_ids = []
        amount_mismatches = []  # { order_number, receivable, csv_net_amount, total_amount, advance_amount, delivery_charge, tax_amount }

        for r in rows:
            order_num = r["order_number"]
            order = order_number_to_order.get(order_num)
            if not order:
                unmatched_order_numbers.append(order_num)
                continue
            matched_order_numbers.append(order_num)
            update_data = {
                "delivery_charge": r["delivery_charge"],
                "tax_amount": r["tax_amount"],
                "courier": "PostEx",
                "updated_at": datetime.utcnow().isoformat(),
            }
            if r.get("tracking_number"):
                update_data["tracking_number"] = r["tracking_number"]
            supabase.table("orders").update(update_data).eq("id", order["id"]).execute()
            updated_order_ids.append(order["id"])
            updated_count += 1

            # Receivable must match grid formula: returned -> -delivery_charge; else -> total - advance - delivery - tax
            total_amount = float(order.get("total_amount") or 0)
            advance_amount = float(order.get("advance_amount") or 0)
            delivery_charge = float(r["delivery_charge"])
            tax_amount = float(r["tax_amount"])
            order_status = (order.get("order_status") or "").strip().lower()
            if order_status == "returned":
                receivable = -delivery_charge
            else:
                receivable = total_amount - advance_amount - delivery_charge - tax_amount
            csv_net = r.get("csv_net_amount")
            if csv_net is not None:
                if round(receivable, 2) != round(float(csv_net), 2):
                    amount_mismatches.append({
                        "order_number": order_num,
                        "receivable": round(receivable, 2),
                        "csv_net_amount": round(float(csv_net), 2),
                        "total_amount": total_amount,
                        "advance_amount": advance_amount,
                        "delivery_charge": delivery_charge,
                        "tax_amount": tax_amount,
                        "order_status": order_status or None,
                    })

        # Build response message with debugging info
        message = f"Updated delivery charges, tax, courier (PostEx), and tracking for {updated_count} order(s)."
        if unmatched_order_numbers:
            message += f" {len(unmatched_order_numbers)} order number(s) from CSV did not match any orders."
            if len(unmatched_order_numbers) <= 10:
                message += f" Unmatched: {', '.join(map(str, unmatched_order_numbers[:10]))}"

        return {
            "updated": updated_count,
            "message": message,
            "updated_order_ids": updated_order_ids,
            "matched_order_numbers": matched_order_numbers,
            "csv_rows_processed": len(rows),
            "csv_order_numbers_count": len(csv_order_numbers),
            "db_order_numbers_count": len(set(db_order_numbers)),
            "matched_count": len(matched_order_numbers),
            "unmatched_count": len(unmatched_order_numbers),
            "amount_mismatches": amount_mismatches,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing CSV: {str(e)}")


@router.get("/{order_id}")
async def get_order(order_id: str):
    """Get a single order by ID"""
    try:
        supabase = get_supabase()
        response = supabase.table("orders").select("*").eq("id", order_id).single().execute()
        if not response.data:
            raise HTTPException(status_code=404, detail="Order not found")
        return response.data
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _delivery_status_is_final(delivery_status_data: dict) -> bool:
    """True if latest status is final (Delivered to Customer or Returned at Merchant Warehouse). No need to fetch from PostEx again."""
    if not delivery_status_data:
        return False
    latest = (delivery_status_data.get("latest_status") or "").strip()
    if not latest:
        return False
    if "Delivered to Customer" in latest:
        return True
    if "Returned at Merchant Warehouse" in latest:
        return True
    return False


def _delivery_status_indicates_returned(delivery_status_data: dict) -> bool:
    """True if delivery status contains 'Return to KARACHI' (e.g. in latest_status or status_history)."""
    if not delivery_status_data:
        return False
    needle = "Return to KARACHI"
    latest = (delivery_status_data.get("latest_status") or "").strip()
    if needle in latest:
        return True
    for item in delivery_status_data.get("status_history") or []:
        if needle in (item.get("status") or ""):
            return True
    return False


def _delivery_status_indicates_delivered(delivery_status_data: dict) -> bool:
    """True if delivery status contains 'Delivered to Customer' (e.g. in latest_status or status_history)."""
    if not delivery_status_data:
        return False
    needle = "Delivered to Customer"
    latest = (delivery_status_data.get("latest_status") or "").strip()
    if needle in latest:
        return True
    for item in delivery_status_data.get("status_history") or []:
        if needle in (item.get("status") or ""):
            return True
    return False


def _delivery_status_indicates_rfd(delivery_status_data: dict) -> bool:
    """True if delivery status contains 'Attempt Made: RFD' (e.g. in latest_status or status_history)."""
    if not delivery_status_data:
        return False
    needle = "Attempt Made: RFD"
    latest = (delivery_status_data.get("latest_status") or "").strip()
    if needle in latest:
        return True
    for item in delivery_status_data.get("status_history") or []:
        if needle in (item.get("status") or ""):
            return True
    return False


def _delivery_status_indicates_ica(delivery_status_data: dict) -> bool:
    """True if delivery status contains 'Attempt Made: ICA' (e.g. in latest_status or status_history)."""
    if not delivery_status_data:
        return False
    needle = "Attempt Made: ICA"
    latest = (delivery_status_data.get("latest_status") or "").strip()
    if needle in latest:
        return True
    for item in delivery_status_data.get("status_history") or []:
        if needle in (item.get("status") or ""):
            return True
    return False


def _delivery_status_indicates_cna(delivery_status_data: dict) -> bool:
    """True if delivery status contains 'Attempt Made: CNA' (e.g. in latest_status or status_history)."""
    if not delivery_status_data:
        return False
    needle = "Attempt Made: CNA"
    latest = (delivery_status_data.get("latest_status") or "").strip()
    if needle in latest:
        return True
    for item in delivery_status_data.get("status_history") or []:
        if needle in (item.get("status") or ""):
            return True
    return False


@router.post("/", response_model=dict)
async def create_order(order: OrderCreate):
    """Create a new order"""
    try:
        supabase = get_supabase()
        order_data = order.model_dump()
        order_data["piece_received"] = "Pending"
        now = datetime.utcnow().isoformat()
        order_data["created_at"] = now
        order_data["updated_at"] = now
        if not order_data.get("order_receiving_date"):
            order_data["order_receiving_date"] = now
        response = supabase.table("orders").insert(order_data).execute()
        return response.data[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def _delivery_status_with_latest_status(existing: Optional[Dict[str, Any]], order_status: str) -> Dict[str, Any]:
    """Build delivery_status JSONB with latest_status set for bulk 'delivered', 'returned', or 'cancelled'.
    Preserves existing keys and uses only JSON-serializable values (str, list, dict).
    """
    if order_status == "delivered":
        latest_status = "Delivered to Customer"
    elif order_status == "returned":
        latest_status = "Return to KARACHI"
    elif order_status == "cancelled":
        latest_status = "Cancelled"
    else:
        latest_status = ""
    now_iso = datetime.utcnow().isoformat()
    # Start from existing JSONB, keeping only JSON-serializable values
    data: Dict[str, Any] = {}
    if existing:
        for k, v in existing.items():
            if k in ("status_history",):
                continue  # we rebuild below
            if v is None or isinstance(v, (str, int, float, bool)):
                data[k] = v
            elif isinstance(v, dict):
                data[k] = {str(a): b for a, b in v.items() if isinstance(b, (str, int, float, bool, type(None)))}
            else:
                data[k] = v
    data["latest_status"] = latest_status
    # status_history: list of { "status": str, "datetime": str } (and optional status_code, is_active)
    history_raw = (existing or {}).get("status_history")
    history: List[Dict[str, Any]] = []
    if isinstance(history_raw, list):
        for item in history_raw:
            if isinstance(item, dict):
                entry = {
                    "status": str(item.get("status", "")),
                    "datetime": str(item.get("datetime", "")),
                }
                if "status_code" in item and isinstance(item["status_code"], str):
                    entry["status_code"] = item["status_code"]
                if "is_active" in item:
                    entry["is_active"] = bool(item["is_active"])
                history.append(entry)
    new_entry: Dict[str, Any] = {"status": latest_status, "datetime": now_iso}
    if history and history[0].get("status") == latest_status:
        pass  # already at front
    else:
        history.insert(0, new_entry)
    data["status_history"] = history
    data["fetched_at"] = now_iso
    return data


class BulkUpdateStatusBody(BaseModel):
    order_numbers: List[int]
    order_status: str  # "delivered", "returned", or "cancelled"


@router.post("/bulk-update-status")
async def bulk_update_order_status(body: BulkUpdateStatusBody):
    """Update order_status and delivery_status.latest_status for multiple orders by order_number."""
    if body.order_status not in ("delivered", "returned", "cancelled"):
        raise HTTPException(status_code=400, detail="order_status must be 'delivered', 'returned', or 'cancelled'")
    if not body.order_numbers:
        raise HTTPException(status_code=400, detail="order_numbers cannot be empty")
    try:
        supabase = get_supabase()
        # Fetch orders so we can merge delivery_status per order
        response = (
            supabase.table("orders")
            .select("id, order_number, delivery_status")
            .in_("order_number", body.order_numbers)
            .execute()
        )
        orders = response.data or []
        updated_at = datetime.utcnow().isoformat()
        updated_order_numbers = []
        for order in orders:
            order_id = order.get("id")
            if not order_id:
                continue
            new_delivery_status = _delivery_status_with_latest_status(
                order.get("delivery_status"), body.order_status
            )
            supabase.table("orders").update({
                "order_status": body.order_status,
                "delivery_status": new_delivery_status,
                "updated_at": updated_at,
            }).eq("id", order_id).execute()
            onum = order.get("order_number")
            if onum is not None:
                updated_order_numbers.append(int(onum))
        requested_set = set(body.order_numbers)
        updated_set = set(updated_order_numbers)
        not_found_order_numbers = sorted(requested_set - updated_set)
        return {
            "updated_count": len(updated_order_numbers),
            "order_status": body.order_status,
            "requested_count": len(body.order_numbers),
            "updated_order_numbers": sorted(updated_order_numbers),
            "not_found_order_numbers": not_found_order_numbers,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class BulkUpdatePieceReceivedBody(BaseModel):
    order_numbers: List[int]


@router.post("/bulk-update-piece-received")
async def bulk_update_piece_received(body: BulkUpdatePieceReceivedBody):
    """Set piece_received to 'Received' for multiple orders by order_number."""
    if not body.order_numbers:
        raise HTTPException(status_code=400, detail="order_numbers cannot be empty")
    try:
        supabase = get_supabase()
        update_data = {
            "piece_received": "Received",
            "updated_at": datetime.utcnow().isoformat(),
        }
        response = (
            supabase.table("orders")
            .update(update_data)
            .in_("order_number", body.order_numbers)
            .execute()
        )
        updated_rows = response.data or []
        updated_order_numbers = [int(o["order_number"]) for o in updated_rows if o.get("order_number") is not None]
        requested_set = set(body.order_numbers)
        updated_set = set(updated_order_numbers)
        not_found_order_numbers = sorted(requested_set - updated_set)
        return {
            "updated_count": len(updated_order_numbers),
            "piece_received": "Received",
            "requested_count": len(body.order_numbers),
            "updated_order_numbers": sorted(updated_order_numbers),
            "not_found_order_numbers": not_found_order_numbers,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{order_id}")
async def update_order(order_id: str, order: OrderUpdate):
    """Update an existing order"""
    try:
        supabase = get_supabase()
        update_data = {k: v for k, v in order.model_dump().items() if v is not None}
        # piece_received defaults to Pending; set to Done when order is delivered (e.g. from delivery-status save)
        update_data["updated_at"] = datetime.utcnow().isoformat()
        response = supabase.table("orders").update(update_data).eq("id", order_id).execute()
        if not response.data:
            raise HTTPException(status_code=404, detail="Order not found")
        return response.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{order_id}/delivery-status")
async def get_delivery_status(order_id: str, save: bool = Query(False, description="If true, store fetched status in order.delivery_status")):
    """Fetch delivery status from courier API. Optionally store in order.delivery_status when save=true."""
    try:
        supabase = get_supabase()
        # Use limit(1) instead of single() so "not found" returns 404, not 500
        order_response = supabase.table("orders").select("*").eq("id", order_id).limit(1).execute()
        
        if not order_response.data or len(order_response.data) == 0:
            raise HTTPException(status_code=404, detail="Order not found")
        
        order = order_response.data[0]
        courier = order.get("courier", "").strip()
        tracking_number = order.get("tracking_number", "").strip()
        
        if not tracking_number:
            raise HTTPException(status_code=400, detail="Tracking number not available")
        
        if courier.upper() == "UNASSIGNED":
            raise HTTPException(status_code=400, detail="Courier not assigned")
        
        delivery_status_data = None
        existing_delivery = order.get("delivery_status")

        if courier.upper() == "POSTEX":
            # If last status is final, no need to call PostEx; return stored details
            if existing_delivery and _delivery_status_is_final(existing_delivery):
                delivery_status_data = existing_delivery
            else:
                async with httpx.AsyncClient(timeout=30.0) as client:
                    api_url = f"https://api.postex.pk/services/courier/api/guest/get-order/{tracking_number}"
                    response = await client.get(api_url)
                    response.raise_for_status()
                    data = response.json()
                    if data.get("statusCode") == "200" and "dist" in data:
                        dist = data["dist"]
                        status_history = dist.get("transactionStatusHistory", [])
                        delivery_status_data = {
                            "courier": "PostEx",
                            "tracking_number": dist.get("trackingNumber", tracking_number),
                            "customer_name": dist.get("customerName", ""),
                            "order_pickup_date": dist.get("orderPickupDate", ""),
                            "status_history": [
                                {
                                    "status": item.get("transactionStatusMessage", ""),
                                    "status_code": item.get("transactionStatusMessageCode", ""),
                                    "datetime": item.get("modifiedDatetime", "")
                                }
                                for item in status_history
                            ],
                            "latest_status": status_history[0].get("transactionStatusMessage", "") if status_history else "",
                            "fetched_at": datetime.utcnow().isoformat()
                        }
        else:
            raise HTTPException(status_code=400, detail="Only PostEx is supported for delivery status tracking")
        
        if not delivery_status_data:
            raise HTTPException(status_code=500, detail="Failed to fetch delivery status")

        # Only persist when we fetched new data (not when we returned stored final status)
        if save:
            update_payload = {
                "delivery_status": delivery_status_data,
                "updated_at": datetime.utcnow().isoformat()
            }
            # Check delivery status and update order_status accordingly
            # Priority: Return > Delivered > RFD > ICA > CNA
            if _delivery_status_indicates_returned(delivery_status_data):
                update_payload["order_status"] = "returned"
            elif _delivery_status_indicates_delivered(delivery_status_data):
                update_payload["order_status"] = "delivered"
                update_payload["piece_received"] = "Done"
            elif _delivery_status_indicates_rfd(delivery_status_data):
                update_payload["order_status"] = "RFD"
            elif _delivery_status_indicates_ica(delivery_status_data):
                update_payload["order_status"] = "ICA"
            elif _delivery_status_indicates_cna(delivery_status_data):
                update_payload["order_status"] = "CNA"
            supabase.table("orders").update(update_payload).eq("id", order_id).execute()

        return delivery_status_data
        
    except httpx.HTTPStatusError as e:
        raise HTTPException(status_code=e.response.status_code, detail=f"Failed to fetch delivery status: {e.response.text}")
    except httpx.RequestError as e:
        err_msg = str(e) or getattr(e, "message", "") or type(e).__name__
        raise HTTPException(
            status_code=500,
            detail=f"Could not reach the courier tracking site ({type(e).__name__}: {err_msg}). "
                   "Check that this server can access the internet and that the courier site is not blocked."
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching delivery status: {str(e)}")

@router.delete("/{order_id}")
async def delete_order(order_id: str):
    """Delete an order"""
    try:
        supabase = get_supabase()
        response = supabase.table("orders").delete().eq("id", order_id).execute()
        return {"message": "Order deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/generate-invoice")
async def generate_invoice(order_ids: List[str]):
    """Generate an Excel invoice from template for selected orders"""
    try:
        if not order_ids:
            raise HTTPException(status_code=400, detail="No orders selected")
        
        # Get orders from database
        supabase = get_supabase()
        orders_response = supabase.table("orders").select("*").in_("id", order_ids).execute()
        orders = orders_response.data
        
        if not orders:
            raise HTTPException(status_code=404, detail="No orders found")
        
        # Load template
        template_path = Path(__file__).parent.parent / "invoice_template.xlsx"
        if not template_path.exists():
            raise HTTPException(status_code=404, detail="Invoice template not found")
        
        # Load workbook with images preserved
        # keep_links=True preserves external links and images
        wb = load_workbook(template_path, keep_links=True)
        ws = wb.active
        
        # Add date to E2 (preserve style) - display format DD/MM/YYYY
        from datetime import datetime
        now = datetime.now()
        current_date = now.strftime("%d/%m/%Y")
        cell_e2 = ws.cell(row=2, column=5)
        if cell_e2.has_style:
            original_style = copy(cell_e2._style)
            cell_e2.value = current_date
            cell_e2._style = original_style
        else:
            cell_e2.value = current_date
        
        # Start filling data from row 7
        start_row = 7
        current_row = start_row
        net_amount_sum = 0
        
        # Store template row styles (row 7) to copy to data rows
        template_row = start_row
        template_styles = {}
        for col in range(1, 6):  # Columns A to E
            template_cell = ws.cell(row=template_row, column=col)
            if template_cell.has_style:
                template_styles[col] = copy(template_cell._style)
        
        # Sort orders by order_number for consistent output
        orders.sort(key=lambda x: x.get("order_number", 0))
        
        for order in orders:
            order_number = order.get("order_number", "")
            tracking_number = order.get("tracking_number", "") or ""
            total_amount = float(order.get("total_amount", 0) or 0)
            advance_amount = float(order.get("advance_amount", 0) or 0)
            delivery_charge = float(order.get("delivery_charge", 0) or 0)
            tax_amount = float(order.get("tax_amount", 0) or 0)
            
            # Calculate COD (total_amount - advance_amount)
            cod = total_amount - advance_amount
            
            # Calculate Net Amount (receivable = total_amount - advance_amount - delivery_charge - tax_amount)
            net_amount = total_amount - advance_amount - delivery_charge - tax_amount
            
            # Fill row data while preserving styles
            values = [order_number, tracking_number, cod, delivery_charge, net_amount]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                # Preserve existing style or use template style
                if cell.has_style:
                    original_style = copy(cell._style)
                    cell.value = value
                    cell._style = original_style
                elif col_idx in template_styles:
                    cell.value = value
                    cell._style = copy(template_styles[col_idx])
                else:
                    cell.value = value
            
            net_amount_sum += net_amount
            current_row += 1
        
        # Add final balance row (preserve styles)
        final_row = current_row
        cell_d = ws.cell(row=final_row, column=4)
        if cell_d.has_style:
            original_style_d = copy(cell_d._style)
            cell_d.value = "Final Balance"
            cell_d._style = original_style_d
            # Make it bold
            if cell_d.font:
                cell_d.font = Font(bold=True, name=cell_d.font.name, size=cell_d.font.size)
            else:
                cell_d.font = Font(bold=True)
        else:
            cell_d.value = "Final Balance"
            cell_d.font = Font(bold=True)
        
        cell_e = ws.cell(row=final_row, column=5)
        if cell_e.has_style:
            original_style_e = copy(cell_e._style)
            cell_e.value = net_amount_sum
            cell_e._style = original_style_e
        else:
            cell_e.value = net_amount_sum
        
        # Configure page setup for printing (first page only)
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = 9  # A4 paper size (9 is the constant for A4)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = 100
        
        # Save to BytesIO buffer for Excel
        from io import BytesIO
        import tempfile
        import os
        
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # PDF conversion disabled for now - return Excel file directly
        # Convert to PDF using win32com (Windows only, requires Excel installed)
        # pdf_buffer = BytesIO()
        # try:
        #     import win32com.client
        #     import pythoncom
        #     
        #     # Create temporary Excel file
        #     with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_excel:
        #         temp_excel.write(excel_buffer.getvalue())
        #         temp_excel_path = temp_excel.name
        #     
        #     # Create temporary PDF file path
        #     with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_pdf:
        #         temp_pdf_path = temp_pdf.name
        #     
        #     # Initialize COM
        #     pythoncom.CoInitialize()
        #     excel = win32com.client.Dispatch("Excel.Application")
        #     excel.Visible = False
        #     excel.DisplayAlerts = False
        #     
        #     try:
        #         # Open the Excel file
        #         workbook = excel.Workbooks.Open(temp_excel_path)
        #         worksheet = workbook.Worksheets(1)  # First worksheet
        #         
        #         # Set print area to first page only
        #         worksheet.PageSetup.PrintArea = ""
        #         worksheet.PageSetup.FitToPagesWide = 1
        #         worksheet.PageSetup.FitToPagesTall = 1
        #         
        #         # Export to PDF (print first page)
        #         worksheet.ExportAsFixedFormat(
        #             Type=0,  # xlTypePDF
        #             Filename=temp_pdf_path,
        #             Quality=0,  # xlQualityStandard
        #             IncludeDocProperties=True,
        #             IgnorePrintAreas=False,
        #             OpenAfterPublish=False
        #         )
        #         
        #         workbook.Close(SaveChanges=False)
        #         
        #         # Read PDF into buffer
        #         with open(temp_pdf_path, 'rb') as pdf_file:
        #             pdf_buffer.write(pdf_file.read())
        #         pdf_buffer.seek(0)
        #         
        #     finally:
        #         excel.Quit()
        #         pythoncom.CoUninitialize()
        #         # Clean up temp files
        #         try:
        #             os.unlink(temp_excel_path)
        #             os.unlink(temp_pdf_path)
        #         except:
        #             pass
        #     
        #     # Return PDF file as download
        #     return Response(
        #         content=pdf_buffer.getvalue(),
        #         media_type="application/pdf",
        #         headers={"Content-Disposition": "attachment; filename=invoice.pdf"}
        #     )
        #     
        # except ImportError:
        #     # Fallback: return Excel file if win32com is not available
        #     return Response(
        #         content=excel_buffer.getvalue(),
        #         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        #         headers={"Content-Disposition": "attachment; filename=invoice.xlsx"}
        #     )
        # except Exception as e:
        #     # If PDF conversion fails, return Excel file
        #     return Response(
        #         content=excel_buffer.getvalue(),
        #         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        #         headers={"Content-Disposition": "attachment; filename=invoice.xlsx"}
        #     )
        
        # Return Excel file directly
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=invoice.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating invoice: {str(e)}")

